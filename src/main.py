"""FastAPI веб-приложение для сравнения PUML диаграмм с JSON результатами OCR."""

import json
from io import BytesIO
from typing import Annotated, Dict, List, Optional

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl.styles import Alignment, Font, PatternFill

from .comparator import compare_puml_with_json, compare_puml_with_puml
from .config import BASE_DIR, MAX_UPLOAD_SIZE
from .services import (
    auto_pair_files,
    normalize_attribute_scores,
    save_uploaded_file,
    session_manager,
)

app = FastAPI(title="PUML vs JSON Comparator")

# Подключение статических файлов и шаблонов
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """Главная страница с формой загрузки"""
    return templates.TemplateResponse(
        "index.html", {"request": request, "max_upload_size": MAX_UPLOAD_SIZE}
    )


@app.post("/preview", response_class=HTMLResponse)
async def preview_upload(
    request: Request,
    puml_files: List[UploadFile] = File(...),
    json_files: List[UploadFile] = File(...),
):
    """Handle initial upload and present pairing preview before comparison."""
    session_id = session_manager.init_session()

    try:
        stored_puml = [
            record
            for upload in puml_files
            if (record := save_uploaded_file(upload, ".puml", session_id))
        ]
        stored_json = [
            record
            for upload in json_files
            if (record := save_uploaded_file(upload, ".json", session_id))
        ]
    except HTTPException as exc:
        session_manager.cleanup_session(session_id)
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": exc.detail,
                "session_id": session_id,
            },
            status_code=exc.status_code,
        )

    if not stored_puml or not stored_json:
        session_manager.cleanup_session(session_id)
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Необходимо загрузить как минимум один PUML и один JSON файл",
                "session_id": session_id,
            },
        )

    pairings, unmatched_json = auto_pair_files(stored_puml, stored_json)
    summary = {
        "puml_count": len(stored_puml),
        "json_count": len(stored_json),
        "paired_count": sum(1 for pair in pairings if pair["json"]),
        "pending_count": sum(1 for pair in pairings if not pair["json"]),
    }

    return templates.TemplateResponse(
        "preview.html",
        {
            "request": request,
            "pairings": pairings,
            "json_options": stored_json,
            "unmatched_json": unmatched_json,
            "summary": summary,
            "session_id": session_id,
        },
    )


@app.post("/upload", response_class=HTMLResponse)
async def run_comparison(
    request: Request,
    puml_files: Annotated[List[str], Form(...)],
    json_choices: Annotated[List[str], Form(...)],
    session_id: Annotated[str, Form(...)],
):
    """Execute comparison based on confirmed pairings from the preview screen."""
    if not puml_files:
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Сессия загрузки не найдена. Пожалуйста, загрузите файлы заново.",
                "session_id": session_id,
            },
        )

    manifest_snapshot = session_manager.get_manifest(session_id)
    if not manifest_snapshot:
        session_manager.cleanup_session(session_id)
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Сессия загрузки не найдена. Пожалуйста, загрузите файлы заново.",
                "session_id": session_id,
            },
        )

    comparisons = []
    for idx, puml_name in enumerate(puml_files):
        json_name = json_choices[idx] if idx < len(json_choices) else ""
        puml_label = manifest_snapshot.get(puml_name, puml_name)
        try:
            puml_path = session_manager.resolve_path(session_id, puml_name)
        except FileNotFoundError:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": "—",
                    "error": "PUML файл не найден. Загрузите диаграммы заново.",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )
            continue

        if not json_name:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": "Не выбран",
                    "error": "Для этого PUML файла не выбрано соответствие",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )
            continue

        json_label = manifest_snapshot.get(json_name, json_name)
        try:
            json_path = session_manager.resolve_path(session_id, json_name)
        except FileNotFoundError:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": json_label,
                    "error": "Указанный JSON файл не найден",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )
            continue

        try:
            result = compare_puml_with_json(str(puml_path), str(json_path))
            result["etalon_file"] = puml_label
            result["student_file"] = json_label
            comparisons.append(result)
        except Exception as e:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": json_label,
                    "error": f"Ошибка при сравнении: {str(e)}",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )

    results = normalize_attribute_scores(comparisons)

    total_score = sum(
        r.get("score", 0) for r in results if "error" not in r or r.get("score", 0) > 0
    )
    avg_score = total_score / len(results) if results else 0

    chart_labels = [r.get("etalon_file", "") for r in results]
    chart_scores = [r.get("score", 0) for r in results]

    stats = {
        "total_comparisons": len(results),
        "avg_score": avg_score,
        "max_score": max(chart_scores) if chart_scores else 0,
        "min_score": min(chart_scores) if chart_scores else 0,
        "passed": sum(1 for s in chart_scores if s >= 70),
        "failed": sum(1 for s in chart_scores if s < 70),
    }

    payload = {
        "results": results,
        "stats": stats,
        "chart_labels": chart_labels,
        "chart_scores": chart_scores,
    }
    session_manager.store_results(session_id, payload)

    return RedirectResponse(
        url=f"/results/{session_id}", status_code=status.HTTP_303_SEE_OTHER
    )


@app.get("/results/{session_id}", response_class=HTMLResponse)
async def show_results(request: Request, session_id: str):
    payload = session_manager.load_results(session_id)
    if not payload:
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Результаты не найдены. Пожалуйста, выполните сравнение заново.",
                "session_id": session_id,
            },
            status_code=status.HTTP_404_NOT_FOUND,
        )

    return templates.TemplateResponse(
        "results.html",
        {
            "request": request,
            "results": payload["results"],
            "stats": payload["stats"],
            "chart_labels": payload["chart_labels"],
            "chart_scores": payload["chart_scores"],
            "session_id": session_id,
        },
    )


@app.get("/export/{session_id}")
async def export_results(session_id: str):
    payload = session_manager.load_results(session_id)
    if not payload:
        raise HTTPException(status_code=404, detail="Session not found")

    results = payload["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    headers = [
        "№",
        "Эталон (PUML)",
        "Распознанный JSON",
        "Оценка",
        "Классы F1",
        "Классы Precision",
        "Классы Recall",
        "Классы (Эталон)",
        "Классы (Студент)",
        "Классы (Совпало)",
        "Атрибуты F1",
        "Атрибуты Precision",
        "Атрибуты Recall",
        "Атрибуты (Эталон)",
        "Атрибуты (Студент)",
        "Атрибуты (Совпало)",
        "Классы: Отсутствуют",
        "Классы: Лишние",
        "Атрибуты: Различия",
        "Ошибка",
    ]

    ws.append(headers)

    # Style header
    header_fill = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, result in enumerate(results, 1):
        classes = result.get("classes", {})
        attributes = result.get("attributes", {})
        diff = result.get("diff", {})
        class_diff = diff.get("classes", {})
        attr_diff = diff.get("attributes", [])

        missing_classes = ", ".join(class_diff.get("missing", []) or [])
        extra_classes = ", ".join(class_diff.get("extra", []) or [])

        attr_diff_parts = []
        for entry in attr_diff:
            parts = []
            name = entry.get("etalon_class") or entry.get("student_class") or "?"
            if entry.get("missing"):
                parts.append(f"нет: {', '.join(entry['missing'])}")
            if entry.get("extra"):
                parts.append(f"лишние: {', '.join(entry['extra'])}")
            if parts:
                attr_diff_parts.append(f"{name} [{' | '.join(parts)}]")
        attr_diff_str = "; ".join(attr_diff_parts)

        row = [
            idx,
            result.get("etalon_file", ""),
            result.get("student_file", ""),
            f"{result.get('score', 0):.1f}%",
            f"{classes.get('f1', 0):.3f}" if classes else "",
            f"{classes.get('precision', 0):.3f}" if classes else "",
            f"{classes.get('recall', 0):.3f}" if classes else "",
            classes.get("etalon_count", "") if classes else "",
            classes.get("student_count", "") if classes else "",
            classes.get("matched", "") if classes else "",
            f"{attributes.get('f1', 0):.3f}" if attributes else "",
            f"{attributes.get('precision', 0):.3f}" if attributes else "",
            f"{attributes.get('recall', 0):.3f}" if attributes else "",
            attributes.get("etalon_count", "") if attributes else "",
            attributes.get("student_count", "") if attributes else "",
            attributes.get("matched", "") if attributes else "",
            missing_classes,
            extra_classes,
            attr_diff_str,
            result.get("error", ""),
        ]
        ws.append(row)

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(
            length + 2, 50
        )  # Cap width at 50

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="puml_comparison_results.xlsx"'
    }
    return StreamingResponse(
        output,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/health")
async def health_check():
    """Проверка здоровья приложения"""
    return {"status": "ok"}


def read_xlsx_metrics(file_content: bytes) -> Dict[str, Dict[str, float]]:
    """
    Reads metrics from an XLSX file.
    Returns a dict: { "Etalon Name": { "score": float, "f1_cls": float, "f1_attr": float } }
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active

    metrics = {}

    # Assuming header is row 1
    # Columns (1-based):
    # 2: Etalon (PUML)
    # 4: Score (e.g. "95.5%")
    # 5: Classes F1
    # 11: Attributes F1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 11:
            continue

        etalon = row[1]
        if not etalon:
            continue

        score_str = str(row[3]) if row[3] is not None else "0"
        try:
            score = float(score_str.replace("%", "").strip()) if score_str else 0.0
        except ValueError:
            score = 0.0

        try:
            f1_cls = float(row[4]) if row[4] is not None and row[4] != "" else 0.0
        except ValueError:
            f1_cls = 0.0

        try:
            f1_attr = float(row[10]) if row[10] is not None and row[10] != "" else 0.0
        except ValueError:
            f1_attr = 0.0

        metrics[etalon] = {
            "score": score,
            "f1_cls": f1_cls,
            "f1_attr": f1_attr,
        }

    return metrics


@app.get("/compare", response_class=HTMLResponse)
async def compare_page(request: Request):
    return templates.TemplateResponse("compare.html", {"request": request})


@app.post("/compare", response_class=HTMLResponse)
async def process_comparison(
    request: Request,
    file_1: Optional[UploadFile] = File(None),
    file_2: Optional[UploadFile] = File(None),
    file_3: Optional[UploadFile] = File(None),
    name_1: str = Form("Подход 1"),
    name_2: str = Form("Подход 2"),
    name_3: str = Form("Подход 3"),
):
    # Filter out empty uploads
    raw_inputs = [
        (file_1, name_1),
        (file_2, name_2),
        (file_3, name_3),
    ]
    valid_inputs = [(f, n) for f, n in raw_inputs if f and f.filename]

    if not valid_inputs:
        return templates.TemplateResponse(
            "compare.html",
            {
                "request": request,
                "error": "Необходимо загрузить хотя бы один файл",
            },
        )

    files = [f for f, n in valid_inputs]
    names = [n for f, n in valid_inputs]

    all_metrics = []

    for file in files:
        content = await file.read()
        try:
            metrics = read_xlsx_metrics(content)
        except Exception:
            # Handle invalid files gracefully-ish
            metrics = {}
        all_metrics.append(metrics)

    # Aggregate data
    # Find all unique diagrams
    all_diagrams = set()
    for m in all_metrics:
        all_diagrams.update(m.keys())

    sorted_diagrams = sorted(list(all_diagrams))

    comparison_rows = []
    for diagram in sorted_diagrams:
        scores = []
        for m in all_metrics:
            scores.append(m.get(diagram, {}).get("score"))

        comparison_rows.append({"diagram": diagram, "scores": scores})

    # Calculate averages
    avg_scores = []
    avg_f1_classes = []
    avg_f1_attributes = []

    for m in all_metrics:
        if not m:
            avg_scores.append(0)
            avg_f1_classes.append(0)
            avg_f1_attributes.append(0)
            continue

        scores = [v["score"] for v in m.values()]
        f1_cls = [v["f1_cls"] for v in m.values()]
        f1_attr = [v["f1_attr"] for v in m.values()]

        avg_scores.append(sum(scores) / len(scores))
        avg_f1_classes.append(sum(f1_cls) / len(f1_cls))
        avg_f1_attributes.append(sum(f1_attr) / len(f1_attr))

    return templates.TemplateResponse(
        "compare_results.html",
        {
            "request": request,
            "approach_names": names,
            "comparison_rows": comparison_rows,
            "avg_scores": avg_scores,
            "avg_f1_classes": avg_f1_classes,
            "avg_f1_attributes": avg_f1_attributes,
        },
    )


@app.post("/preview-puml-puml", response_class=HTMLResponse)
async def preview_puml_puml(
    request: Request,
    puml_files: List[UploadFile] = File(...),
    puml2_files: List[UploadFile] = File(...),
):
    """Handle initial upload for PUML vs PUML mode."""
    session_id = session_manager.init_session()

    try:
        stored_puml1 = [
            record
            for upload in puml_files
            if (record := save_uploaded_file(upload, ".puml", session_id))
        ]
        stored_puml2 = [
            record
            for upload in puml2_files
            if (record := save_uploaded_file(upload, ".puml", session_id))
        ]
    except HTTPException as exc:
        session_manager.cleanup_session(session_id)
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": exc.detail,
                "session_id": session_id,
            },
            status_code=exc.status_code,
        )

    if not stored_puml1 or not stored_puml2:
        session_manager.cleanup_session(session_id)
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Необходимо загрузить файлы в оба набора",
                "session_id": session_id,
            },
        )

    # Автоматическое заполнение пар: каждому файлу из первого набора подбираем файл из второго
    # Если файлы совпадают по имени - автоматически, иначе назначаем первый доступный
    pairings = []
    used_puml2 = set()

    for puml1 in stored_puml1:
        match = None
        # Сначала ищем точное совпадение по имени
        for puml2 in stored_puml2:
            if puml2["filename"] in used_puml2:
                continue
            if puml1["label"].split(".")[0] == puml2["label"].split(".")[0]:
                match = puml2
                used_puml2.add(puml2["filename"])
                break

        # Если точного совпадения нет, берем первый доступный
        if not match:
            for puml2 in stored_puml2:
                if puml2["filename"] not in used_puml2:
                    match = puml2
                    used_puml2.add(puml2["filename"])
                    break

        pairings.append({"puml": puml1, "json": match})

    unmatched_puml2 = [
        entry for entry in stored_puml2 if entry["filename"] not in used_puml2
    ]

    summary = {
        "puml_count": len(stored_puml1),
        "json_count": len(stored_puml2),
        "paired_count": sum(1 for pair in pairings if pair["json"]),
        "pending_count": sum(1 for pair in pairings if not pair["json"]),
    }

    return templates.TemplateResponse(
        "preview.html",
        {
            "request": request,
            "pairings": pairings,
            "json_options": stored_puml2,
            "unmatched_json": unmatched_puml2,
            "summary": summary,
            "session_id": session_id,
            "mode": "puml-puml",
        },
    )


@app.post("/upload-puml-puml", response_class=HTMLResponse)
async def run_comparison_puml_puml(
    request: Request,
    puml_files: Annotated[List[str], Form(...)],
    json_choices: Annotated[List[str], Form(...)],
    session_id: Annotated[str, Form(...)],
):
    """Execute PUML vs PUML comparison."""
    if not puml_files:
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Сессия загрузки не найдена. Пожалуйста, загрузите файлы заново.",
                "session_id": session_id,
            },
        )

    manifest_snapshot = session_manager.get_manifest(session_id)
    if not manifest_snapshot:
        session_manager.cleanup_session(session_id)
        return templates.TemplateResponse(
            "results.html",
            {
                "request": request,
                "error": "Сессия загрузки не найдена. Пожалуйста, загрузите файлы заново.",
                "session_id": session_id,
            },
        )

    comparisons = []
    for idx, puml_name in enumerate(puml_files):
        puml2_name = json_choices[idx] if idx < len(json_choices) else ""
        puml_label = manifest_snapshot.get(puml_name, puml_name)
        try:
            puml_path = session_manager.resolve_path(session_id, puml_name)
        except FileNotFoundError:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": "—",
                    "error": "PUML файл не найден. Загрузите диаграммы заново.",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )
            continue

        if not puml2_name:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": "Не выбран",
                    "error": "Для этого PUML файла не выбрано соответствие",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )
            continue

        puml2_label = manifest_snapshot.get(puml2_name, puml2_name)
        try:
            puml2_path = session_manager.resolve_path(session_id, puml2_name)
        except FileNotFoundError:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": puml2_label,
                    "error": "Указанный PUML файл не найден",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )
            continue

        try:
            result = compare_puml_with_puml(str(puml_path), str(puml2_path))
            result["etalon_file"] = puml_label
            result["student_file"] = puml2_label
            comparisons.append(result)
        except Exception as e:
            comparisons.append(
                {
                    "etalon_file": puml_label,
                    "student_file": puml2_label,
                    "error": f"Ошибка при сравнении: {str(e)}",
                    "similarity": 0.0,
                    "score": 0.0,
                }
            )

    results = normalize_attribute_scores(comparisons)

    total_score = sum(
        r.get("score", 0) for r in results if "error" not in r or r.get("score", 0) > 0
    )
    avg_score = total_score / len(results) if results else 0

    chart_labels = [r.get("etalon_file", "") for r in results]
    chart_scores = [r.get("score", 0) for r in results]

    stats = {
        "total_comparisons": len(results),
        "avg_score": avg_score,
        "max_score": max(chart_scores) if chart_scores else 0,
        "min_score": min(chart_scores) if chart_scores else 0,
        "passed": sum(1 for s in chart_scores if s >= 70),
        "failed": sum(1 for s in chart_scores if s < 70),
    }

    payload = {
        "results": results,
        "stats": stats,
        "chart_labels": chart_labels,
        "chart_scores": chart_scores,
    }
    session_manager.store_results(session_id, payload)

    return RedirectResponse(
        url=f"/results/{session_id}", status_code=status.HTTP_303_SEE_OTHER
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
